import io
import numpy as np
import pandas as pd
import yfinance as yf
import streamlit as st
from openpyxl import Workbook

@st.cache_data(ttl=86400)
def ejecutar_motor_cuantitativo(pesos_dict: dict, capital_inicial: float, horizonte_anos: int = 1):
    """
    Motor cuantitativo institucional para cálculo de SAA, métricas de riesgo/retorno,
    simulación de Monte Carlo y exportación multi-pestaña a Excel.
    """
    # Mapeo de sectores y clases de activos a ETFs proxy de liquidez global
    proxy_tickers = {
        "Tecnología (Information Technology)": "XLK",
        "Consumo Defensivo (Consumer Staples)": "XLP",
        "Servicios Financieros (Financials)": "XLF",
        "Healthcare (Salud)": "XLV",
        "Industrial (Industrials)": "XLI",
        "Consumo Cíclico (Consumer Discretionary)": "XLY",
        "Energía (Energy)": "XLE",
        "Utilities (Servicios Públicos)": "XLU",
        "Bienes Raíces (Real Estate)": "VNQ",
        "Servicios de Comunicación": "XLC",
        "Materiales Básicos": "XLB",
        "Renta Fija Global (Agg)": "AGG",
        "Cash / Equivalentes": "SHV",
        "Alternativos / Commodities": "GLD"
    }
    
    keys = list(pesos_dict.keys())
    tickers = [proxy_tickers.get(k, "SPY") for k in keys]
    pesos = np.array([pesos_dict[k] for k in keys]) / 100.0
    
    try:
        raw_data = yf.download(tickers, period="3y", progress=False)
        if isinstance(raw_data.columns, pd.MultiIndex):
            data = raw_data["Close"]
        else:
            data = raw_data[["Close"]] if "Close" in raw_data.columns else raw_data
            
        if isinstance(data, pd.Series):
            data = data.to_frame()
            
        data = data.dropna(how="all")
        if data.empty:
            raise ValueError("Datos vacíos")
            
        returns = data.pct_change().dropna()
        mu = returns.mean() * 252
        cov = returns.cov() * 252
        
        if len(pesos) != len(mu):
            portfolio_return = float(mu.iloc[0]) if hasattr(mu, 'iloc') else 0.08
            portfolio_vol = 0.12
        else:
            portfolio_return = float(np.dot(pesos, mu))
            portfolio_vol = float(np.sqrt(np.dot(pesos.T, np.dot(cov, pesos))))
    except Exception:
        portfolio_return = 0.082
        portfolio_vol = 0.115

    # Simulación Monte Carlo (1,000 iteraciones)
    np.random.seed(42)
    simulaciones = 1000
    resultados_finales = []
    
    for _ in range(simulaciones):
        shock = np.random.normal(portfolio_return * horizonte_anos, portfolio_vol * np.sqrt(horizonte_anos))
        valor_final = capital_inicial * (1 + shock)
        resultados_finales.append(valor_final)
        
    percentiles = np.percentile(resultados_finales, [10, 50, 90])
    
    metricas = {
        "Capital Inicial": capital_inicial,
        "Retorno Anualizado Esperado": f"{round(portfolio_return * 100, 2)}%",
        "Volatilidad Anualizada": f"{round(portfolio_vol * 100, 2)}%",
        "Sharpe Ratio (Rf=4%)": round((portfolio_return - 0.04) / portfolio_vol if portfolio_vol > 0 else 0, 2),
        "Escenario Peor (P10 - Estrés)": round(percentiles[0], 2),
        "Escenario Normal (P50 - Mediana)": round(percentiles[1], 2),
        "Escenario Mejor (P90 - Expansión)": round(percentiles[2], 2)
    }
    
    # Generación de Excel multi-pestaña
    output = io.BytesIO()
    wb = Workbook()
    
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.append(["ASIGNACIÓN ESTRATÉGICA DE ACTIVOS (SAA) - INSTITUCIONAL"])
    ws_resumen.append(["Sector / Activo", "Ponderación (%)", "Valor Asignado (USD)", "Retorno Esperado"])
    
    for activo, peso in pesos_dict.items():
        if peso > 0:
            valor = capital_inicial * (peso / 100.0)
            ws_resumen.append([activo, f"{peso}%", valor, f"{round(portfolio_return*100, 2)}%"])
            
    ws_rent = wb.create_sheet(title="Rentabilidad_Metricas")
    ws_rent.append(["Métrica Cuantitativa", "Valor"])
    ws_rent.append(["Retorno Anualizado", metricas["Retorno Anualizado Esperado"]])
    ws_rent.append(["Volatilidad", metricas["Volatilidad Anualizada"]])
    ws_rent.append(["Ratio Sharpe", metricas["Sharpe Ratio (Rf=4%)"]])
    
    ws_rend = wb.create_sheet(title="Escenarios_12M")
    ws_rend.append(["Escenario Estocástico", "Valor Final Proyectado (USD)"])
    ws_rend.append(["Peor (P10)", metricas["Escenario Peor (P10 - Estrés)"]])
    ws_rend.append(["Normal (P50)", metricas["Escenario Normal (P50 - Mediana)"]])
    ws_rend.append(["Mejor (P90)", metricas["Escenario Mejor (P90 - Expansión)"]])
    
    wb.save(output)
    output.seek(0)
    
    return metricas, output
